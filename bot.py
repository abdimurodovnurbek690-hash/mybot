import os
import telebot
from telebot import types
from openpyxl import Workbook, load_workbook

TOKEN = os.getenv("BOT_TOKEN")
bot = telebot.TeleBot(TOKEN)

ADMIN_ID = 2133751835  # 👉 BU YERGA O'ZINGNI ID

user_data = {}

# 💰 NARXLAR
prices = {
    "🍎 Olma": 10000,
    "🍌 Banan": 15000
}

# 📊 EXCEL FUNKSIYA
def save_to_excel(product, price, phone, address):
    file = "orders.xlsx"

    if not os.path.exists(file):
        wb = Workbook()
        ws = wb.active
        ws.append(["Mahsulot", "Narx", "Telefon", "Manzil"])
        wb.save(file)

    wb = load_workbook(file)
    ws = wb.active
    ws.append([product, price, phone, address])
    wb.save(file)

# 🚀 START
@bot.message_handler(commands=['start'])
def start(message):
    markup = types.ReplyKeyboardMarkup(resize_keyboard=True)
    markup.add("📦 Buyurtma berish")
    bot.send_message(message.chat.id, "Assalomu alaykum! Buyurtma berish uchun tugmani bosing.", reply_markup=markup)

# 📦 BUYURTMA BOSHLASH
@bot.message_handler(func=lambda m: m.text == "📦 Buyurtma berish")
def order(message):
    markup = types.ReplyKeyboardMarkup(resize_keyboard=True)
    markup.add("🍎 Olma", "🍌 Banan")
    bot.send_message(message.chat.id, "Mahsulotni tanlang:", reply_markup=markup)

# 🛒 MAHSULOT TANLASH
@bot.message_handler(func=lambda m: m.text in prices)
def product(message):
    user_data[message.chat.id] = {"product": message.text}

    markup = types.ReplyKeyboardMarkup(resize_keyboard=True)
    btn = types.KeyboardButton("📞 Telefon yuborish", request_contact=True)
    markup.add(btn)

    bot.send_message(message.chat.id, "Telefon raqamingizni yuboring:", reply_markup=markup)

# 📞 TELEFON
@bot.message_handler(content_types=['contact'])
def contact(message):
    user_data[message.chat.id]["phone"] = message.contact.phone_number
    bot.send_message(message.chat.id, "📍 Manzilingizni yozing:")

# 📍 MANZIL
@bot.message_handler(func=lambda message: True)
def address(message):
    data = user_data.get(message.chat.id, {})

    product = data.get("product", "Noma'lum")
    phone = data.get("phone", "Noma'lum")
    address = message.text
    price = prices.get(product, 0)

    # 📦 FOYDALANUVCHIGA
    bot.send_message(message.chat.id,
        f"📦 Yangi buyurtma:\n"
        f"{product}\n"
        f"💰 {price} so'm\n"
        f"📞 {phone}\n"
        f"📍 {address}"
    )

    bot.send_message(message.chat.id, "✅ Buyurtma qabul qilindi!")

    # 💳 TO‘LOV
    markup = types.InlineKeyboardMarkup()
    btn = types.InlineKeyboardButton("💳 To‘lov qilish", url="https://payme.uz")
    markup.add(btn)

    bot.send_message(message.chat.id, "💰 To‘lovni amalga oshiring:", reply_markup=markup)

    # 👨‍💻 ADMIN
    bot.send_message(ADMIN_ID,
        f"🆕 BUYURTMA\n\n"
        f"📦 {product}\n"
        f"💰 {price} so'm\n"
        f"📞 {phone}\n"
        f"📍 {address}"
    )

    # 📊 EXCEL
    save_to_excel(product, price, phone, address)

bot.infinity_polling()
